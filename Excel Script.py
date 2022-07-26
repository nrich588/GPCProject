
import time
from tkinter.tix import CELL
from openpyxl import load_workbook
import openpyxl

# this function provides a basic delayText and print option, allowing both tasts to be accomplished with a single line of code. A delayText after the message can be included as well.
def delayText(dur1,message="",dur2=0):
    time.sleep(dur1)
    print(message)
    time.sleep(dur2)

def viewSheet():
    sheetPath = "book.xlsx"
    delayText(1,"Displaying Data...")


def editSheet():
    
    #loads the proper book
    bookPath = "book.xlsx"
    book = load_workbook(bookPath)
    delayText(1,"Edit mode entered.")
    
    # this loop makes sure the user selects an actual sheet to work in
    while True:
        delayText(0.5,"Please enter the name of the sheet you would like to open.")
        #this prints the sheetnames by loading the workbook and then its sheets.
        delayText(0.5,book.sheetnames,0.5)
        selectedSheet = input()
        # if the enter sheetname is in the book.sheetnames array, it will attempt to open it.
        if selectedSheet in book.sheetnames:
            delayText(0.5,"Opening " + selectedSheet + "sheet.")
            loadedSheet = book[selectedSheet]
            #iter_rows() only iterates through rows/cells with data, need way to find the next row with column 1 empty.
            for row in loadedSheet.iter_rows():
                for cell in row:
                    if cell.value == None:
                        startPoint = [cell.row, cell.column]
                        print(startPoint)
                        break
                    else:
                        print("No empty cells?")

        else:
            delayText(0.5,"Please enter a valid sheet name!")
    
    
    
    
    
    
    
    


def main():
    
    delayText(1,"Initializing GPC Excel Utility.")
    while True:    
        delayText(1,"Would you like to edit or view the purchase history spreadsheet?")
        delayText(1,"1.) Edit     2.) View     3.) Exit")
        editOrView = input()
        editOrView = editOrView.lower()
        #ask if they want to add or view information
        if editOrView == "1":
            editSheet()
            # if add then run add function
        elif editOrView == "2":
            viewSheet()
            # if view then run view function
        elif editOrView == "3":
            break
        else:
            delayText(1,"Please enter a valid input.")


main()


# notes, it is important to remember that excel has a specific structure of BOOK -> SHEET -> CELL